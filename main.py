import asyncio
import contextlib
import io
import json
import logging
import os
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import pandas as pd
from aiogram import Bot, Dispatcher, Router
from aiogram.enums import ChatMemberStatus, ChatType
from aiogram.exceptions import AiogramError, TelegramBadRequest
from aiogram.filters import Command, Filter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    Message,
    PollAnswer,
    Update
)
from dotenv import load_dotenv
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from scheduler import Scheduler
from storage import StorageManager


# ----- Configuration -----
@dataclass(frozen=True)
class Config:
    token: str
    chat_id: int
    group_id: int
    admin_ids: list[int]
    test_mode: bool
    poll_interval: float = 60.0
    prefetch_offset: float = 300.0
    poll_window: float = 300.0
    include_exams: bool = False
    nmg_types: list[str] = field(default_factory=list)

    @staticmethod
    def _parse_admin_ids(raw: str) -> list[int]:
        try:
            return [int(x.strip()) for x in raw.split(',') if x.strip()]
        except ValueError:
            return []

    @staticmethod
    def _convert_to_bool(value: str) -> bool:
        return value.strip().lower() in ('true', '1', 't', 'y', 'yes')

    @classmethod
    def from_env(cls) -> 'Config':
        load_dotenv()
        token = os.getenv('TOKEN', '')
        if not token or token == 'token':
            raise RuntimeError('Environment variable TOKEN is invalid.')

        chat_id = int(os.getenv('CHAT_ID', '0'))
        group_id = int(os.getenv('GROUP_ID', '0'))
        admin_ids = cls._parse_admin_ids(os.getenv('ADMIN_COMMANDS_ACCESS'))
        test_mode = cls._convert_to_bool(os.getenv('TEST_MODE', 'false'))
        poll_interval = float(os.getenv('POLL_CHECK_INTERVAL', '60'))
        poll_window = float(os.getenv('POLL_CLOSURE_WINDOW', '300'))
        prefetch_offset = float(os.getenv('SCHEDULE_PREFETCH_OFFSET', '300'))
        include_exams = cls._convert_to_bool(os.getenv('INCLUDE_EXAMS', 'false'))
        nmg_types = os.getenv('NMG_TYPES', 'practice,lab').split(',')

        return cls(
            token=token,
            chat_id=chat_id,
            group_id=group_id,
            admin_ids=admin_ids,
            test_mode=test_mode,
            poll_interval=poll_interval,
            poll_window=poll_window,
            prefetch_offset=prefetch_offset,
            include_exams=include_exams,
            nmg_types=nmg_types
        )


# ----- Logging -----
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ----- Validators -----
NAME_REGEX = re.compile(r'^[А-Яа-яЁё-]+$')
DISCIPLINE_QUOTE = re.compile(r'"([^"]+)"')


def is_valid_name(name: str) -> bool:
    return bool(NAME_REGEX.fullmatch(name.strip()))


def extract_quoted(text: str) -> list[str]:
    return re.findall(r'"([^"]+)"', text)


async def is_valid_chat_type(bot: Bot, config: Config) -> None:
    try:
        chat = await bot.get_chat(config.chat_id)
        if chat.type not in (ChatType.GROUP, ChatType.SUPERGROUP):
            raise RuntimeError(
                f'Only groups and supergroups are supported. Invalid chat type: {chat.type}, ID: {chat.id}'
            )
    except AiogramError as e:
        raise RuntimeError(f'Could not fetch chat info: {e}')

    return None


class UserInGroupFilter(Filter):
    def __init__(self, bot: Bot, chat_id: int):
        self.bot = bot
        self.chat_id = chat_id

    async def __call__(self, message: Message) -> bool:
        try:
            member = await self.bot.get_chat_member(self.chat_id, message.from_user.id)
        except TelegramBadRequest as e:
            logging.error(f'Error when fetching chat member info: {e}')
            return False

        return member.status in {
            ChatMemberStatus.CREATOR,
            ChatMemberStatus.ADMINISTRATOR,
            ChatMemberStatus.MEMBER,
            ChatMemberStatus.RESTRICTED,
        }


class PrivateChatFilter(Filter):
    async def __call__(self, message: Message) -> bool:
        if message.chat.type != ChatType.PRIVATE:
            await message.answer('Эта команда доступна только в личных сообщениях.')
            return False
        return True


class AdminFilter(Filter):
    def __init__(self, admin_ids: list[int]):
        self.admin_ids = admin_ids

    async def __call__(self, message: Message) -> bool:
        if message.from_user.id not in self.admin_ids:
            await message.answer('Эта команда доступна только администраторам.')
            return False
        return True


# ----- FSM States -----
class Registration(StatesGroup):
    last_name = State()
    first_name = State()


class ManageDisciplineState(StatesGroup):
    action = State()
    full_class_name = State()
    alias = State()
    class_type = State()


class ManageEmojiBanState(StatesGroup):
    action = State()
    emoji = State()


# ----- Main Bot Class -----
class AttendanceBot:
    def __init__(self, config: Config):
        self.config = config
        self.bot: Bot | None = None
        self.dispatcher: Dispatcher | None = None
        self.router = Router()
        self.storage = StorageManager()
        self.scheduler: Scheduler | None = None

    def setup_routes(self) -> None:
        group_filter = UserInGroupFilter(self.bot, self.config.chat_id)
        dm_filter = PrivateChatFilter()
        admin_filter = AdminFilter(self.config.admin_ids)

        self.router.message(Command('start'), group_filter)(self._on_start)
        self.router.message(Command('edit_name'), dm_filter, group_filter)(self._on_edit_name)
        self.router.message(Command('display_name'), dm_filter, group_filter)(self._on_display_name)
        self.router.message(Command('export_attendance'), dm_filter, group_filter)(self._on_export_attendance)
        self.router.message(Command('check_permissions'), group_filter, admin_filter)(self._on_check_permissions)

        self.router.message(
            Command('manage_disciplines'),
            dm_filter,
            group_filter,
            admin_filter
        )(self._on_manage_disciplines_menu)
        self.router.callback_query(lambda c: c.data and c.data.startswith('md:'))(self._on_manage_discipline_cb)
        self.router.message(ManageDisciplineState.full_class_name)(self._on_receive_full_class_name)
        self.router.message(ManageDisciplineState.alias)(self._on_receive_class_alias)
        self.router.message(ManageDisciplineState.class_type)(self._on_receive_class_type)

        self.router.message(
            Command('manage_emoji_bans'),
            dm_filter,
            group_filter,
            admin_filter
        )(self._on_manage_emoji_bans_menu)
        self.router.callback_query(lambda c: c.data == 'eb:list')(self._on_show_banned_emojis_list)
        self.router.callback_query(lambda c: c.data and c.data.startswith('eb:') and c.data != 'eb:list')(self._on_manage_emoji_ban_cb)
        self.router.message(ManageEmojiBanState.emoji)(self._on_receive_emoji)
        self.router.message(Registration.last_name)(self._on_last_name)
        self.router.message(Registration.first_name)(self._on_first_name)
        # Message moderation - must be last to not interfere with commands
        self.router.message(lambda m: m.chat.type in (ChatType.GROUP, ChatType.SUPERGROUP))(self._on_group_message)
        self.router.poll_answer()(self._on_poll_answer)
        self.router.errors()(self._on_error)

    async def _build_report(self, polls: list[dict[str, Any]], year: int, month: int) -> BufferedInputFile:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:  # type: ignore
            df = pd.DataFrame(polls)
            for date_val, group in df.groupby('date'):
                sheet_name = str(date_val).replace('.', '-')[:31]

                label_to_time: dict[str, str] = {}
                records: dict[str, dict[str, str]] = {}

                for _, row in group.iterrows():
                    cls_label = f'{row["class_name"]} ({row["start_time"]} - {row["end_time"]})'

                    if cls_label in label_to_time:
                        prof_last_name = row['prof'].split()[1]
                        cls_label = f'{cls_label} ({prof_last_name})'

                    label_to_time[cls_label] = row["start_time"]

                    for resp in json.loads(row['responses']):
                        user = await self.storage.get_user(resp['user_id']) or {}
                        last_name = user.get('last_name', resp['last_name'])
                        first_name = user.get('first_name', resp['first_name'])

                        name = f'{last_name} {first_name}'
                        opt = resp['option_ids'][0] if resp['option_ids'] else None
                        mark = {0: 'Д', 1: 'Н', 2: 'П', 3: 'Б', 4: 'НМГ'}.get(opt, '')
                        records.setdefault(name, {})[cls_label] = mark

                classes_list = sorted(
                    label_to_time.keys(),
                    key=lambda lbl: datetime.strptime(label_to_time[lbl], "%H:%M")
                )

                table = []
                for student_name, answers in records.items():
                    safe_name = student_name
                    if isinstance(student_name, str) and student_name[0] in ('=', '+', '-', '@'):
                        safe_name = f"'{student_name}"  # excel injection protection

                    row_dict = {'Имя': safe_name}
                    for cls_label in classes_list:
                        row_dict[cls_label] = answers.get(cls_label, '')
                    table.append(row_dict)

                df_day = pd.DataFrame(table, columns=['Имя'] + classes_list)
                df_day = df_day.sort_values(by='Имя')
                df_day.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(df_day.columns, start=1):
                    max_width = max(
                        df_day[col].astype(str).map(len).max(),
                        len(str(col))
                    )
                    adjusted_width = max_width + 3
                    col_letter = get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = adjusted_width

                align = Alignment(horizontal='center', vertical='center')
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = align

        output.seek(0)
        return BufferedInputFile(output.getvalue(), filename=f'attendance_{year}-{month:02d}.xlsx')

    # ----- Route Handlers -----
    async def _on_start(self, message: Message, state: FSMContext) -> None:
        await state.clear()
        user_id = str(message.from_user.id)
        user = await self.storage.get_user(user_id)

        if message.chat.type == ChatType.PRIVATE and not (user and user.get('registered')):
            await message.answer('Введите вашу фамилию (одно русское слово):')
            await state.set_state(Registration.last_name)
        else:
            start_time = time.monotonic()
            response = await message.answer('Бот запущен!')
            ping = int((time.monotonic() - start_time) * 1000)

            await response.edit_text(
                f'Бот запущен!\n'
                f'Пинг: {ping} мс\n'
                f'(Жду следующей пары...)'
            )

    async def _on_edit_name(self, message: Message, state: FSMContext) -> None:
        await state.clear()
        await message.answer('Введите вашу фамилию (одно русское слово):')
        return await state.set_state(Registration.last_name)

    async def _on_display_name(self, message: Message, state: FSMContext) -> Message:
        await state.clear()
        user_id = str(message.from_user.id)
        user = await self.storage.get_user(user_id) or {}

        if not user:
            return await message.answer('Ваше имя не указано.')

        last = user.get('last_name')
        first = user.get('first_name')

        return await message.answer(f'Ваше имя: {last} {first}.')

    async def _on_check_permissions(self, message: Message) -> Message:
        try:
            bot_member = await self.bot.get_chat_member(message.chat.id, self.bot.id)
            status = bot_member.status
            can_delete = status in (ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR)

            banned_emojis = await self.storage.get_banned_emojis()
            emoji_count = len(banned_emojis)

            return await message.answer(
                f'🔍 Проверка прав бота:\n'
                f'Статус бота: {status}\n'
                f'Может удалять сообщения: {"✅ Да" if can_delete else "❌ Нет"}\n'
                f'Запрещённых смайликов: {emoji_count}\n'
                f'Список: {" ".join(banned_emojis) if banned_emojis else "пуст"}'
            )
        except Exception as e:
            return await message.answer(f'❌ Ошибка при проверке прав: {e}')

    async def _on_last_name(self, message: Message, state: FSMContext) -> Message | None:
        last = message.text.capitalize()
        if not is_valid_name(last):
            return await message.answer(
                'Фамилия должна состоять из одного русского слова. Пожалуйста, введите корректную фамилию:'
            )

        await state.update_data(last_name=last)
        await message.answer('Отлично! Теперь введите ваше имя (одно русское слово):')

        return await state.set_state(Registration.first_name)

    async def _on_first_name(self, message: Message, state: FSMContext) -> Message | None:
        first = message.text.capitalize()
        if not is_valid_name(first):
            return await message.answer(
                'Имя должно состоять из одного русского слова. Пожалуйста, введите корректное имя:'
            )

        data = await state.get_data()
        user_id = str(message.from_user.id)

        await self.storage.update_user(user_id, {
            'username': message.from_user.username or '',
            'last_name': data['last_name'],
            'first_name': first,
            'registered': True
        })

        await message.answer(f'Спасибо, {data["last_name"]} {first}! Ваши данные сохранены.')
        return await state.clear()

    async def _on_export_attendance(self, message: Message, state: FSMContext) -> Message | None:
        await state.clear()

        parts = message.text.split(maxsplit=1)
        now = datetime.now()
        year, month = now.year, now.month

        if len(parts) > 1:
            try:
                year, month = map(int, parts[1].split('-', maxsplit=1))
            except ValueError:
                return await message.answer('Используйте формат: /export_attendance YYYY-MM')

        await message.answer(f'Генерирую отчёт за {year}-{month:02d}…')

        polls = await self.storage.get_past_polls_by_month(year, month)
        if not polls:
            return await message.answer('Нет данных за этот период.')

        file = await self._build_report(polls, year, month)
        return await message.answer_document(file)

    async def _on_manage_disciplines_menu(self, message: Message) -> Message:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text='✏️ Сократить', callback_data='md:set_alias'),
                InlineKeyboardButton(text='👥 Подгруппы', callback_data='md:set_nmg'),  # not my group / не моя группа
            ],
            [
                InlineKeyboardButton(text='🗑️ Исключить', callback_data='md:exclude'),
            ],
        ])
        return await message.answer('Выберите действие для дисциплин:', reply_markup=kb)

    @staticmethod
    async def _on_manage_discipline_cb(query: CallbackQuery, state: FSMContext) -> Message | None:
        await query.answer()
        action = query.data.split(':', maxsplit=1)[1]  # set_alias, set_nmg, exclude
        await state.update_data(action=action)
        class_name_query = ('Введите полное название дисциплины в кавычках,'
                            'например: "Введение в математический анализ"')

        if action == 'set_alias':
            await query.message.answer(
                f'Эта функция позволяет сократить название дисциплины в опросах и таблицах.\n\n'
                f'{class_name_query}'
            )
        if action == 'set_nmg':
            await query.message.answer(
                f'Эта функция позволяет добавить опцию "Не моя группа"'
                f'в опросах для дисциплины с указанным типом занятия.\n\n'
                f'{class_name_query}'
            )
        if action == 'exclude':
            await query.message.answer(
                f'Эта функция позволяет исключить дисциплину из опросов.\n\n'
                f'{class_name_query}'
            )

        return await state.set_state(ManageDisciplineState.full_class_name)

    async def _on_receive_full_class_name(self, message: Message, state: FSMContext) -> Message | None:
        parts = extract_quoted(message.text)
        if len(parts) != 1:
            return await message.answer('Нужно ровно одно название в кавычках.')
        full = parts[0]

        data = await state.get_data()
        action = data['action']
        if not action:
            await state.clear()
            return await message.answer('Ошибка, начните заново через /manage_disciplines.')

        await state.update_data(full_name=full)

        if action == 'set_alias':
            await message.answer(f'Теперь введите сокращение для "{full}" в кавычках:')
            await state.set_state(ManageDisciplineState.alias)

        if action == 'set_nmg':
            formatted_nmg_types = str(self.config.nmg_types).strip('[]')
            await message.answer(
                f'Теперь введите тип занятия для "{full}" в кавычках.'
                f'Допустимые значения: {formatted_nmg_types}'
            )
            await state.set_state(ManageDisciplineState.class_type)

        if action == 'exclude':
            await self.storage.set_discipline_setting(full, is_excluded=True)
            await state.clear()
            logger.info(f'Excluded: {full}')
            return await message.answer(f'Дисциплина "{full}" добавлена в список исключённых.')

        return None

    async def _on_receive_class_type(self, message: Message, state: FSMContext):
        data = await state.get_data()
        full = data.get('full_name')
        if not full:
            await state.clear()
            return await message.answer('Ошибка, начните заново через /manage_disciplines.')

        class_type = extract_quoted(message.text.strip())[0]
        if not class_type:
            return await message.answer('Тип занятия не может быть пустым.')

        await self.storage.set_discipline_setting(full, class_type=class_type, is_nmg=True)
        await state.clear()
        logger.info(f'Added NMG: {full} ({class_type})')
        return await message.answer(f'Добавлено в список "не моя группа": {full} ({class_type})')

    async def _on_receive_class_alias(self, message: Message, state: FSMContext):
        data = await state.get_data()
        full = data.get('full_name')
        if not full:
            await state.clear()
            return await message.answer('Ошибка, начните заново через /manage_disciplines.')

        alias = extract_quoted(message.text.strip())[0]
        if not alias:
            return await message.answer('Короткое название не может быть пустым.')

        await self.storage.set_discipline_setting(full, alias=alias)
        await state.clear()
        logger.info(f'Added alias: "{full}" → {alias}')
        return await message.answer(f'Добавлено: "{full}" → "{alias}"')

    # ----- Emoji ban management -----
    async def _on_manage_emoji_bans_menu(self, message: Message) -> Message:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text='➕ Добавить в бан', callback_data='eb:add'),
                InlineKeyboardButton(text='➖ Убрать из бана', callback_data='eb:remove'),
            ],
            [
                InlineKeyboardButton(text='📋 Список запрещённых', callback_data='eb:list'),
            ],
        ])
        return await message.answer('Выберите действие для управления запрещёнными смайликами:', reply_markup=kb)

    @staticmethod
    async def _on_manage_emoji_ban_cb(query: CallbackQuery, state: FSMContext) -> Message | None:
        await query.answer()
        action = query.data.split(':', maxsplit=1)[1]  # add, remove, list
        await state.update_data(action=action)

        if action == 'list':
            # Show list without changing state
            return None
        elif action == 'add':
            await query.message.answer(
                'Отправьте смайлик, который хотите добавить в бан-список.\n'
                'Можно отправить несколько смайликов в одном сообщении.'
            )
        elif action == 'remove':
            await query.message.answer(
                'Отправьте смайлик, который хотите убрать из бан-списка.\n'
                'Можно отправить несколько смайликов в одном сообщении.'
            )

        return await state.set_state(ManageEmojiBanState.emoji)

    async def _on_receive_emoji(self, message: Message, state: FSMContext) -> Message | None:
        data = await state.get_data()
        action = data.get('action')
        if not action:
            await state.clear()
            return await message.answer('Ошибка, начните заново через /manage_emoji_bans.')

        text = message.text.strip()
        if not text:
            return await message.answer('Сообщение не может быть пустым.')

        # Extract emojis from text (simple approach - any non-text characters)
        emojis = []
        for char in text:
            if ord(char) > 127:  # Non-ASCII characters are likely emojis
                emojis.append(char)

        if not emojis:
            return await message.answer('В сообщении не найдено смайликов. Попробуйте отправить смайлики.')

        user_id = str(message.from_user.id)
        added_count = 0
        removed_count = 0

        for emoji in emojis:
            if action == 'add':
                if await self.storage.add_banned_emoji(emoji, user_id):
                    added_count += 1
                    logger.info(f'Added banned emoji: {emoji} by {user_id}')
                else:
                    await message.answer(f'Смайлик {emoji} уже в бан-списке.')
            elif action == 'remove':
                if await self.storage.remove_banned_emoji(emoji):
                    removed_count += 1
                    logger.info(f'Removed banned emoji: {emoji} by {user_id}')
                else:
                    await message.answer(f'Смайлик {emoji} не найден в бан-списке.')

        await state.clear()

        if action == 'add':
            return await message.answer(f'Добавлено в бан-список: {added_count} смайликов')
        elif action == 'remove':
            return await message.answer(f'Убрано из бан-списка: {removed_count} смайликов')

    async def _on_show_banned_emojis_list(self, query: CallbackQuery) -> None:
        await query.answer()

        banned_emojis = await self.storage.get_banned_emojis()

        if not banned_emojis:
            text = '📋 Запрещённых смайликов нет.'
        else:
            emoji_list = ' '.join(banned_emojis)
            text = f'📋 Запрещённые смайлики ({len(banned_emojis)}):\n{emoji_list}'

        await query.message.answer(text)

    async def _on_group_message(self, message: Message) -> None:
        # Skip if user is admin
        if message.from_user.id in self.config.admin_ids:
            return

        # Skip if message has no text
        if not message.text:
            return

        # Check if bot has admin rights in this chat
        try:
            bot_member = await self.bot.get_chat_member(message.chat.id, self.bot.id)
            if not bot_member.status in (ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR):
                logger.warning(f'Bot lacks admin rights in chat {message.chat.id}')
                return  # Bot can't delete messages
        except TelegramBadRequest:
            logger.warning(f'Cannot check bot permissions in chat {message.chat.id}')
            return  # Can't check permissions

        # Get banned emojis
        banned_emojis = await self.storage.get_banned_emojis()
        if not banned_emojis:
            return

        # Check if message contains banned emojis
        found_banned_emojis = []
        for char in message.text:
            if char in banned_emojis:
                found_banned_emojis.append(char)

        if not found_banned_emojis:
            return

        logger.info(f'Deleting message with banned emoji(s): {found_banned_emojis} from user {message.from_user.id}')

        # Delete the message
        try:
            await message.delete()
            logger.info(f'Successfully deleted message with banned emoji(s): {found_banned_emojis} from user {message.from_user.id}')

            # Send warning message
            warning_text = f'⚠️ Сообщение пользователя {message.from_user.first_name} удалено за использование запрещённых смайликов: {" ".join(found_banned_emojis)}'
            await message.answer(warning_text)

        except TelegramBadRequest as e:
            logger.error(f'Failed to delete message: {e}')
        except Exception as e:
            logger.error(f'Unexpected error in message moderation: {e}')

    async def _on_poll_answer(self, poll_answer: PollAnswer) -> None:
        user = poll_answer.user

        rec = {
            'poll_id': poll_answer.poll_id,
            'user_id': str(user.id),
            'option_ids': poll_answer.option_ids,
            'first_name': user.first_name,
            'last_name': user.last_name or '',
            'username': f'@{user.username}' if user.username else ''
        }
        await self.storage.update_poll_response(
            rec['poll_id'], rec['user_id'], rec['option_ids'], rec['first_name'], rec['last_name'], rec['username']
        )
        for entry in self.scheduler.active_polls.values():
            if entry['poll_id'] == poll_answer.poll_id:
                responses = json.loads(entry['responses'])
                responses.append(rec)
                entry['responses'] = json.dumps(responses)
                break

    @staticmethod
    async def _on_error(update: Update, exception: Exception | None = None) -> None:
        logger.exception(f'Error for update {update}: {exception!r}')

    async def run(self) -> None:
        await self.storage.connect()
        self.bot = Bot(token=self.config.token)
        self.dispatcher = Dispatcher()
        self.dispatcher.include_router(self.router)
        self.setup_routes()

        try:
            await is_valid_chat_type(self.bot, self.config)
        except RuntimeError as e:
            logger.critical(e)
            return None

        # discipline settings:
        # (excluded from polls, "not my group" poll option, class name aliases)
        # ([class_name], {'class_name': class_type}, {'class_name': alias})
        self.scheduler = Scheduler(
            bot=self.bot,
            config=self.config,
            storage=self.storage,
            discipline_settings=await self.storage.get_discipline_settings()
        )

        scheduler_task = asyncio.create_task(self.scheduler.start())
        try:
            await self.dispatcher.start_polling(self.bot)
        finally:
            scheduler_task.cancel()
            with contextlib.suppress(asyncio.CancelledError):
                await scheduler_task

            await self.scheduler.close()
            await self.bot.session.close()


if __name__ == '__main__':
    try:
        attendance_bot = AttendanceBot(Config.from_env())
        asyncio.run(attendance_bot.run())
    except KeyboardInterrupt:
        logger.info('Bot stopped by user.')
